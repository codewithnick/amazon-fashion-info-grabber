from operator import sub
from unicodedata import category
import openpyxl
import traceback
from selenium import webdriver
import time
import os
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
import json
import random
import urllib.parse as urlparse
from urllib.parse import parse_qs
from urllib.parse import urlencode
from randomintro import Intro
def formaturl(url):
    captured_value=url.split("&")[0]
    return captured_value+"&tag=stuvera-20"
def updateparams(url,page):
    params = {'page':page}

    url_parts = list(urlparse.urlparse(url))
    query = dict(urlparse.parse_qsl(url_parts[4]))
    query.update(params)

    url_parts[4] = urlencode(query)
    return urlparse.urlunparse(url_parts)
def createexcelifnotexists():
    filepath="database.xlsx"
    if not os.path.isfile(filepath):
        #if not exists file
        wb = openpyxl.Workbook()
        wb.save(filepath)
def readfromexcel(sheet_name,row,col):
    createexcelifnotexists()
    wb_obj = openpyxl.load_workbook("database.xlsx")
    try:    
        sheet_obj = wb_obj[sheet_name]
    except KeyError:
        wb_obj.create_sheet(sheet_name)
        sheet_obj = wb_obj[sheet_name]
    return sheet_obj.cell(row = row, column = col).value

def writetoexcel(sheet_name,row,col,value):
    createexcelifnotexists()
    wb_obj = openpyxl.load_workbook("database.xlsx")
    try:    
        sheet_obj = wb_obj[sheet_name]
    except KeyError:
        wb_obj.create_sheet(sheet_name)
        sheet_obj = wb_obj[sheet_name]
    sheet_obj.cell(row = row, column = col).value= value
    wb_obj.save("database.xlsx")
def getmaxrow(sheet_name):
    createexcelifnotexists()
    wb_obj = openpyxl.load_workbook("database.xlsx")
    try:    
        sheet_obj = wb_obj[sheet_name]
    except KeyError:
        wb_obj.create_sheet(sheet_name)
        sheet_obj = wb_obj[sheet_name]
    return sheet_obj.max_row

class Browser:
    def __init__(self,):
        chrome_options = Options()
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument("--disable-logging")
        chrome_options.add_argument('log-level=3')
        prefs = {
            "download_restrictions": 3,
        }
        chrome_options.add_experimental_option(
            "prefs", prefs
        )
        driver = webdriver.Chrome('chromedriver',options=chrome_options)
        self.driver=driver          
        # random intro
        self.intro=Intro()
    def start(self,keyword,page):
        #search page 1
        createexcelifnotexists()
        self.driver.get("https://www.amazon.com/")
        time.sleep(1)
        self.driver.find_element_by_id("twotabsearchtextbox").send_keys(keyword+'\n')
        if(page!=1):
            self.driver.get(updateparams(self.driver.current_url,page))
        link=[]
        i=1

        #checking if scrapable or need to restart
        mytext=[]
        while mytext==[]:
            try:
                mytext=self.driver.find_elements_by_xpath("//a[@class='a-link-normal s-underline-text s-underline-link-text s-link-style a-text-normal']")
                print("links found ",len(mytext))
                if(mytext==[]):
                    mystring="No results for "+keyword
                    mystring=mystring.lower()
                    if mystring in self.driver.find_element_by_tag_name("body").text.lower():
                        print("no results available lets return")
                        return
                    print("unable to fetch links ...retrying")
                    self.leave()
                    self.__init__()
                    self.driver.get("https://www.amazon.com/")
                    self.driver.find_element_by_id("twotabsearchtextbox").send_keys(keyword+'\n')
                    if(page!=1):
                        self.driver.get(updateparams(self.driver.current_url,page))
                    self.driver.execute_script("window.scrollTo(0,6000)")
                    time.sleep(5)
                    mytext=self.driver.find_elements_by_xpath("//a[@class='a-link-normal s-link-style a-text-normal']")
            except:
                continue
        
        #yes scrapable
        for i in mytext:
            link.append(i.get_attribute("href"))
        print("links fetched ",len(link))
        #only first result
        count=0
        for i in range(len(link)):
            if count>=5:
                break
            link[i]=formaturl(link[i])
            self.driver.get(link[i])
            #get title
            
            try:
                title=self.driver.find_element_by_id("title").text
                #print(title)
            except:
                #traceback.print_exc()
                #print("title na")
                title=''
                continue
            try:
                section=self.driver.find_element_by_id("wayfinding-breadcrumbs_feature_div").text
                section=section.split("\n")[-1]
            except:
                #traceback.print_exc()
                section=''
            try:
                image=self.driver.find_element_by_id("landingImage").get_attribute("src")
                #print(image)
            except:                
                image=''
                #traceback.print_exc()
            
            #get product description
            """ try:
                self.driver.find_element_by_xpath("//a[@data-action='a-expander-toggle']").click()
            except:
                pass """
            try:          
                try:
                    self.driver.find_element_by_class_name("a-expander-header a-declarative a-expander-extend-header").click()
                except:
                    pass
                desc1=desc2=desc3=''    
                desc1=self.driver.find_element_by_id("productOverview_feature_div").text
                desc2=self.driver.find_element_by_id("featurebullets_feature_div").text
                time.sleep(5)
                self.driver.execute_script("window.scrollTo(0,3000)")
                desc3=self.driver.find_element_by_id("aplus").text
                #print(desc3)
            except:
                try:
                    desc3=self.driver.find_element_by_id("productDescription").text
                except:
                    try:
                        desc3=self.driver.find_element_by_id("btf-content-1_feature_div").text
                    except:
                        print("desc n/a")
                        continue
            #get product details
            try:
                product=self.driver.find_element_by_id("prodDetails")
                product=product.find_element_by_xpath("./div")
                product= product.text
                #print(product)
            except:
                traceback.print_exc()                
                product=''
            
            #formatting
            if ':' in title:
                title=title.split(":")[0]
            if ';' in title:
                title=title.split(";")[0]
            if ',' in title:
                title=title.split(",")[0]
            
            newrow=getmaxrow("Sheet")+1
            writetoexcel(sheet_name="Sheet",row=newrow,col=1,value=title)
            
            purchase='<a href ="{}" class="dl_button"> Buy Now via Amazon </a>'.format(link[i])
            desctext=self.intro.get_intro(title,section)+"\n\n"+ desc1+desc2+desc3 +"\n\n<h2>Product Information</h2>\n\n"
            desctext+=product+"\n\n"+purchase
            print(desctext)

            writetoexcel(sheet_name="Sheet",row=newrow,col=2,value=desctext)
            writetoexcel(sheet_name="Sheet",row=newrow,col=3,value=section)
            writetoexcel(sheet_name="Sheet",row=newrow,col=4,value=keyword)
            writetoexcel(sheet_name="Sheet",row=newrow,col=5,value=link[i])
            writetoexcel(sheet_name="Sheet",row=newrow,col=6,value=image)
            count+=1
            
    def leave(self):
        print(" killing the browser ")
        self.driver.quit()